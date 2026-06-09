"""
GET /api/gmail/recent — fetch recent emails
POST /api/gmail/extract-events — Gemini analyzes emails and proposes calendar events
POST /api/calendar/create-event — create a calendar event
POST /api/gmail/auto-sync — full pipeline: fetch → extract → return proposals
"""

import asyncio
import io
import json
import re
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from pydantic import BaseModel

from gcal import (
    is_configured,
    list_recent_emails,
    list_recent_emails_all_accounts,
    list_upcoming_events,
    create_event as gcal_create_event,
)
from data_manager import now_jst
from router import call_ai, DEFAULT_MODELS

router = APIRouter()


@router.get("/gmail/status")
def gmail_status():
    """Check if Google integration (Gmail + Calendar write) is configured."""
    return {"configured": is_configured()}


@router.get("/gmail/recent")
def gmail_recent(
    days: int = Query(3, ge=1, le=365),
    limit: int = Query(20, ge=1, le=200),
    slot: int = Query(0, ge=0, le=9),  # 0 = all configured slots, otherwise specific slot
):
    """Fetch recent emails. slot=0 fetches all accounts; slot=N fetches that account only."""
    if not is_configured():
        raise HTTPException(status_code=400, detail="Google integration not configured")
    try:
        if slot == 0:
            emails = list_recent_emails_all_accounts(days=days, max_results=limit)
        else:
            emails = list_recent_emails(days=days, max_results=limit, slot=slot)
        return {"emails": emails, "count": len(emails)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gmail fetch failed: {e}")


def _raise_gcal_error(e: Exception, action: str):
    """Map Google API errors to clearer HTTP responses."""
    msg = str(e)
    if "invalid_grant" in msg or "Token has been expired" in msg or "Token has been revoked" in msg:
        raise HTTPException(
            status_code=401,
            detail="GOOGLE_TOKEN_EXPIRED: Google OAuth トークンが失効しました。ローカルで `python scripts/setup_gcal.py` を実行し、新しい token.json の base64 を Railway の GOOGLE_TOKEN_JSON_B64 に貼り直してください。",
        )
    raise HTTPException(status_code=500, detail=f"{action}: {e}")


@router.get("/calendar/account")
def calendar_account():
    """Return the email address whose primary calendar is being used."""
    if not is_configured():
        raise HTTPException(status_code=400, detail="not configured")
    try:
        from gcal import _get_service
        service = _get_service()
        about = service.calendarList().get(calendarId="primary").execute()
        return {
            "calendar_id": about.get("id", ""),
            "summary": about.get("summary", ""),
            "timezone": about.get("timeZone", ""),
        }
    except HTTPException:
        raise
    except Exception as e:
        _raise_gcal_error(e, "calendar account lookup failed")


@router.get("/gmail/slots")
def gmail_slots():
    """Return list of configured account slots so frontend can chunk fetches per-account."""
    from gcal import _configured_slots
    return {"slots": _configured_slots()}


class ExtractRequest(BaseModel):
    days: int = 3
    limit: int = 20
    engine: str = "gemini"  # default: Gemini for long context
    model: str | None = None

    def model_post_init(self, _ctx):
        if self.days < 1:
            self.days = 1
        if self.days > 365:
            self.days = 365
        if self.limit < 1:
            self.limit = 1
        if self.limit > 200:
            self.limit = 200


class EventProposal(BaseModel):
    title: str
    start_iso: str
    end_iso: str
    description: str = ""
    location: str = ""
    confidence: str = "medium"  # high/medium/low
    event_type: str = "default"  # meeting/committee/deadline/default
    source_email_id: str = ""
    source_subject: str = ""


class ExtractFromEmailsRequest(BaseModel):
    """Accept a pre-fetched batch of emails for analysis. Frontend chunks long-range fetches."""
    emails: list[dict]
    engine: str = "gemini"
    model: str | None = None


BATCH_SIZE = 8  # emails per AI call — small enough that JSON output rarely truncates


def _build_system_prompt(today_str: str) -> str:
    return f"""You extract calendar event candidates from emails for a Hokkaido University
associate professor (志柿). Goal: surface the user's REAL commitments — especially
university work — and DROP marketing so the list stays manageable.

TODAY IS: {today_str}

INCLUDE (extract these):
- 大学事務 (総務 / 教務 / 会計) の連絡: 会議・委員会・教授会・授業/講義・休講補講・試験・
  締切・提出・履修・成績・出張・旅費精算・面談
- 他の大学教員 / 共著者 / 学生 からの 授業・課題・締切・研究・学会・科研費 の連絡
  (送信元が個人アドレス @gmail 等でも、内容が教育・研究・学務なら必ず含める)
- 本人自身の予約・予定: 旅行・ホテル・フライト・医療通院・受験(TOEFL 等)・家族/保育園
- 本人が出席 / 提出 / 対応する必要がある、日時の明示された予定

DROP (これらは出力しない — ノイズ):
- 副業・求人の案内: 案件紹介・事業説明会・座談会・スカウト
  (サンカク / クラウドワークス / リクルート / 「副業」「案件」系)
- セール・割引・クーポン・ポイント・キャンペーン・メルマガ・ニュースレターの販促
- ショッピング / サブスクの宣伝告知 (コストコ / アカチャンホンポ / SUZURI / モンベル 等)
- 決済受領・出金・請求通知のみで本人の行動が不要なもの (Zaim / 各種 receipt)
- ニュース要約 (TLDR / Bloomberg / NewsPicks 等) の本文中に出てくる日付

判断基準: 「本人が出る/出す/対応する予定」または「大学・教育・研究の連絡」なら INCLUDE。
「売り込み・販促・案件紹介」なら、たとえ具体的な日付があっても DROP。

Output rules:
- Output ONLY a JSON array. No markdown, no commentary.
- Each item: {{title, start_iso, end_iso, description, location, confidence, event_type, source_email_id, source_subject}}
- Use ISO 8601 with timezone (e.g., "2026-05-15T14:00:00+09:00") OR all-day "YYYY-MM-DD"
- Default timezone: Asia/Tokyo (+09:00)
- If only an approximate date is mentioned (今週中, 来週など), set confidence=low and pick a reasonable date
- confidence: "high" (explicit date/time), "medium" (inferred), "low" (vague or implicit)
- event_type:
    "meeting" — 会議・打ち合わせ・面談・授業
    "committee" — 委員会・理事会・審議会・教授会
    "deadline" — 締め切り・提出期限・due
    "default" — リマインダー、その他
- title: concise, in the email's language (Japanese stays Japanese)
- If multiple real commitments are in one email, output multiple entries

Return [] if nothing actionable. 迷ったら、学務・本人予定は残し、販促は捨てる。"""


def _process_batch(batch_emails: list, system_prompt: str, engine: str, model: str) -> tuple[list, str | None, str | None]:
    """Synchronous batch processor — runs in thread pool. Returns (normalized_proposals, parse_error, raw_preview)."""
    emails_text = "\n\n---EMAIL---\n".join(
        f"ID: {e['id']} (account {e.get('_account_slot', 1)})\nFrom: {e['from']}\nSubject: {e['subject']}\nDate: {e['date']}\n\n{e['body'] or e['snippet']}"
        for e in batch_emails
    )
    user_msg = f"Extract calendar events from these emails:\n\n{emails_text}"

    try:
        response = call_ai(
            messages=[{"role": "user", "content": user_msg}],
            system=system_prompt,
            engine=engine,
            model=model,
            max_tokens=16000,
        )
    except Exception as e:
        return [], f"AI call failed: {e}", None

    cleaned = response.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)

    # 配列の開始位置に揃える (前置きテキストがあっても拾う)
    first = cleaned.find("[")
    if first > 0:
        cleaned = cleaned[first:]

    parse_error = None
    try:
        proposals = json.loads(cleaned)
        if not isinstance(proposals, list):
            proposals = []
    except json.JSONDecodeError as e:
        # 出力が途中で切れていても、完成済みオブジェクトまで救出する
        # (素の json.loads だけだとバッチ全員 = 約8通分の抽出を取りこぼす)
        parse_error = str(e)
        proposals = []
        repaired = _repair_truncated_json_array(cleaned)
        if repaired is not None:
            try:
                salvaged = json.loads(repaired)
                if isinstance(salvaged, list):
                    proposals = salvaged
                    parse_error = f"recovered_partial: {e} (kept {len(salvaged)})"
            except json.JSONDecodeError:
                proposals = []

    normalized = []
    for p in proposals:
        if not isinstance(p, dict):
            continue
        normalized.append({
            "title": str(p.get("title", ""))[:200],
            "start_iso": str(p.get("start_iso", "")),
            "end_iso": str(p.get("end_iso", "")),
            "description": str(p.get("description", ""))[:500],
            "location": str(p.get("location", ""))[:200],
            "confidence": p.get("confidence", "medium"),
            "event_type": p.get("event_type", "default"),
            "source_email_id": str(p.get("source_email_id", "")),
            "source_subject": str(p.get("source_subject", ""))[:200],
        })

    raw_preview = response[:500] if not normalized else None
    return normalized, parse_error, raw_preview


@router.post("/gmail/extract-events")
async def extract_events(req: ExtractRequest):
    """Use Gemini (or other engine) to extract event candidates from recent emails. Batches in parallel for long ranges."""
    if not is_configured():
        raise HTTPException(status_code=400, detail="Google integration not configured")

    try:
        emails = list_recent_emails_all_accounts(days=req.days, max_results=req.limit)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gmail fetch failed: {e}")

    if not emails:
        return {"proposals": [], "emails_scanned": 0}

    today_str = now_jst().strftime("%Y-%m-%d (%A)")
    system_prompt = _build_system_prompt(today_str)

    engine = req.engine if req.engine in DEFAULT_MODELS else "gemini"
    model = req.model or DEFAULT_MODELS.get(engine, DEFAULT_MODELS["gemini"])

    batches = [emails[i:i + BATCH_SIZE] for i in range(0, len(emails), BATCH_SIZE)]

    loop = asyncio.get_event_loop()
    results = await asyncio.gather(*[
        loop.run_in_executor(None, _process_batch, b, system_prompt, engine, model)
        for b in batches
    ])

    all_proposals: list = []
    parse_errors: list[str] = []
    raw_previews: list[str] = []
    for props, err, preview in results:
        all_proposals.extend(props)
        if err:
            parse_errors.append(err)
        if preview:
            raw_previews.append(preview)

    # Deduplicate by (title, start_iso) — same event mentioned in multiple emails
    seen = set()
    deduped = []
    for p in all_proposals:
        key = (p["title"], p["start_iso"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(p)

    return {
        "proposals": deduped,
        "emails_scanned": len(emails),
        "batches": len(batches),
        "batch_size": BATCH_SIZE,
        "engine_used": engine,
        "model_used": model,
        "ai_raw_preview": raw_previews[0] if raw_previews and not deduped else None,
        "parse_error": "; ".join(parse_errors) if parse_errors else None,
    }


@router.post("/gmail/extract-events-from-emails")
def extract_events_from_emails(req: ExtractFromEmailsRequest):
    """Analyze a pre-fetched batch of emails. Frontend uses this for long-range chunked processing."""
    if not req.emails:
        return {"proposals": [], "emails_scanned": 0}

    today_str = now_jst().strftime("%Y-%m-%d (%A)")
    system_prompt = _build_system_prompt(today_str)
    engine = req.engine if req.engine in DEFAULT_MODELS else "gemini"
    model = req.model or DEFAULT_MODELS.get(engine, DEFAULT_MODELS["gemini"])

    proposals, parse_error, raw_preview = _process_batch(req.emails, system_prompt, engine, model)

    return {
        "proposals": proposals,
        "emails_scanned": len(req.emails),
        "engine_used": engine,
        "model_used": model,
        "ai_raw_preview": raw_preview if not proposals else None,
        "parse_error": parse_error,
    }


def _build_pdf_native_prompt(today_str: str, source_label: str) -> str:
    return f"""この PDF からカレンダーに追加すべき予定をすべて抽出してください。

今日の日付: {today_str}
ファイル名: {source_label}

PDF には大学関連の日程表（教授会・研究科会議・学科会議・委員会・FD/SD・採点会議・卒論審査・入試業務・年間予定表・時間割など）が含まれている可能性が高い。表組みの日程はすべて1件ずつ別エントリとして抽出してください（10件以上ある場合も省略せず全部出す）。

出力ルール:
- JSON 配列のみ出力。マークダウン・コメント・余計な説明は一切なし
- 各要素: {{"title", "start_iso", "end_iso", "description", "location", "confidence", "event_type", "recurrence", "source_email_id", "source_subject"}}
- start_iso/end_iso: ISO 8601 with +09:00 (例 "2026-06-10T14:00:00+09:00") または終日なら "YYYY-MM-DD"
- event_type: "meeting" | "committee" | "deadline" | "default"
  - 教授会・研究科会議・委員会系 → "committee"
  - 普通の会議・打合せ・授業 → "meeting"
  - 締切・提出期限 → "deadline"
  - その他 → "default"
- recurrence: 週次繰り返し（時間割など）の場合のみ "FREQ=WEEKLY" 形式、それ以外は ""
- source_email_id: "" のまま
- source_subject: ファイル名 "{source_label}" を使う
- 年が書かれていない場合は今日の日付から推測（今が5月で「6月10日」とあれば 2026-06-10）
- 終了時刻が無い場合は開始から1時間後を仮置き

日程表に20件あれば20件、50件あれば50件、すべて返す。1件も漏らさず、しかし作り話は禁止（書かれていない日付は出さない）。"""


def _repair_truncated_json_array(s: str) -> str | None:
    """For a truncated JSON array, walk back to the last balanced object and close the array.

    Handles strings/escapes so we don't split inside a value. Returns repaired text or None.
    """
    # Find the opening [
    start = s.find("[")
    if start == -1:
        return None
    depth = 0  # object depth
    in_string = False
    escape = False
    last_complete_end = -1  # index of the '}' that closed an object at depth 0 (i.e. an array element)
    for i in range(start, len(s)):
        c = s[i]
        if in_string:
            if escape:
                escape = False
            elif c == "\\":
                escape = True
            elif c == '"':
                in_string = False
            continue
        if c == '"':
            in_string = True
        elif c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                last_complete_end = i
        elif c == "]" and depth == 0:
            return s[start : i + 1]  # already complete
    if last_complete_end == -1:
        return None
    return s[start : last_complete_end + 1] + "]"


def _parse_proposals_json(raw: str, source_label: str) -> tuple[list[dict], str | None]:
    cleaned = raw.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    if not cleaned.startswith("["):
        first = cleaned.find("[")
        if first != -1:
            cleaned = cleaned[first:]
    # Ensure closing ]
    if not cleaned.rstrip().endswith("]"):
        last = cleaned.rfind("]")
        if last != -1:
            cleaned = cleaned[: last + 1]
    parse_error: str | None = None
    proposals: list = []
    try:
        proposals = json.loads(cleaned)
        if not isinstance(proposals, list):
            proposals = []
    except json.JSONDecodeError as e:
        parse_error = str(e)
        # Repair: walk back to the last well-formed "},"" boundary and close the array there.
        repaired = _repair_truncated_json_array(cleaned)
        if repaired is not None:
            try:
                proposals = json.loads(repaired)
                if isinstance(proposals, list):
                    parse_error = f"recovered_partial: {parse_error} (kept {len(proposals)} items)"
                else:
                    proposals = []
            except json.JSONDecodeError:
                proposals = []
    normalized: list[dict] = []
    for p in proposals:
        if not isinstance(p, dict):
            continue
        normalized.append({
            "title": str(p.get("title", ""))[:200],
            "start_iso": str(p.get("start_iso", "")),
            "end_iso": str(p.get("end_iso", "")),
            "description": str(p.get("description", ""))[:500],
            "location": str(p.get("location", ""))[:200],
            "confidence": p.get("confidence", "medium"),
            "event_type": p.get("event_type", "default"),
            "recurrence": str(p.get("recurrence", "")),
            "source_email_id": "",
            "source_subject": source_label,
        })
    return normalized, parse_error


def _extract_with_ai(full_text: str, source_label: str, engine: str) -> dict:
    """Shared AI extraction for PDF/Excel/text. Returns the same shape as the PDF endpoint."""
    today_str = now_jst().strftime("%Y-%m-%d (%A)")
    system_prompt = f"""You extract calendar events from a document (schedule sheet, syllabus, timetable, meeting agenda).

TODAY IS: {today_str}

The text comes from a PDF/Excel — may contain table-formatted dates, semester schedules, exam schedules, committee schedules, weekly timetables.

Items to extract:
- All meetings, classes, exams, committees, deadlines with dates/times
- 大学関連: 教授会・研究科会議・学科会議・学部会議・研究会・FD・SD・入試業務・採点会議・卒論審査・教務委員会・人事委員会・予算委員会・各種ワーキンググループ etc.
- All-day events if no time specified
- Weekly recurring classes (timetable/時間割): include `recurrence` with the RRULE-compatible info
- Be GENEROUS — include anything resembling a scheduled gathering, even if labeled vaguely (例: ○○について、説明会、報告会、懇談会)

**重要 — 表形式の日程表（教授会日程表・年間予定表など）**:
- 表の行で「6/10(火) 14:00〜16:00 教授会」のような形式を見つけたら、すべて個別のイベントとして抽出
- 月だけの行（例: 「6月」セル → その下に複数の日付）の場合、各日付ごとに別エントリ作成
- 同じ会議が複数月にわたって列挙されている場合（例: 教授会 4/15, 5/13, 6/10, 7/8, …）も全て抽出（必要なら一度に20-50件）
- パイプ区切り `|` で来た行はテーブル行 — 各セルを意味的に解釈（曜日 / 日付 / 時間 / 会議名 / 場所）
- 「定例○○会議」「第N回○○」のような繰り返し型でも、一覧されている全日程を個別エントリにする

Output rules:
- Output ONLY a JSON array. No markdown, no commentary.
- Each item: {{title, start_iso, end_iso, description, location, confidence, event_type, recurrence, source_email_id, source_subject}}
- ISO 8601 with timezone (+09:00) or all-day "YYYY-MM-DD"
- source_email_id: leave ""
- source_subject: use the document filename/sheet name
- event_type: meeting/committee/deadline/default
- title: in Japanese if source is Japanese
- If a year is omitted, assume the most likely year given today's date
- For weekly recurring (e.g. "月曜 1限 線形代数"), set:
    - start_iso to the FIRST occurrence (e.g. semester start week, that Monday)
    - recurrence: "FREQ=WEEKLY;UNTIL=YYYYMMDDT235959Z" (semester end if known, otherwise omit UNTIL)
    - If recurrence not needed (one-off), set recurrence: ""
- For Excel timetable cells, infer the time from period columns (1限=08:45-10:15, 2限=10:30-12:00, 3限=13:00-14:30, 4限=14:45-16:15, 5限=16:30-18:00) unless the document specifies otherwise

Return [] only if no dates can be identified."""

    user_msg = f"Extract calendar events from this document (source: {source_label}):\n\n{full_text[:80000]}"
    selected_engine = engine if engine in DEFAULT_MODELS else "gemini"
    model = DEFAULT_MODELS.get(selected_engine, DEFAULT_MODELS["gemini"])

    try:
        response = call_ai(
            messages=[{"role": "user", "content": user_msg}],
            system=system_prompt,
            engine=selected_engine,
            model=model,
            max_tokens=8192,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI extraction failed: {e}")

    normalized, parse_error = _parse_proposals_json(response, source_label)
    return {
        "proposals": normalized,
        "engine_used": selected_engine,
        "model_used": model,
        "parse_error": parse_error,
        "ai_raw_preview": response[:1500] if not normalized else None,
    }


@router.post("/calendar/extract-events-from-excel")
async def extract_events_from_excel(
    file: UploadFile = File(...),
    engine: str = "gemini",
):
    """Extract calendar candidates from an uploaded Excel (.xlsx) — supports timetables (時間割)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not available")

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="empty file")

    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel parse failed: {e}")

    sheets_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            sheets_text.append(f"### Sheet: {sheet_name}\n" + "\n".join(rows))
    full_text = "\n\n".join(sheets_text)

    if not full_text.strip():
        return {
            "proposals": [],
            "parse_error": "Excel から内容を抽出できませんでした",
            "filename": file.filename,
            "sheet_count": len(wb.sheetnames),
        }

    result = _extract_with_ai(full_text, file.filename or "uploaded.xlsx", engine)
    result["filename"] = file.filename
    result["sheet_count"] = len(wb.sheetnames)
    return result


@router.post("/calendar/extract-events-from-pdf")
async def extract_events_from_pdf(
    file: UploadFile = File(...),
    engine: str = "gemini",
):
    """Extract calendar candidates from an uploaded PDF (日程表・年間予定表 etc.)."""
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="empty file")

    # Path A: send the PDF directly to Gemini (multimodal — handles tables/layout/scanned natively)
    try:
        import google.generativeai as genai
        from data_manager import now_jst as _now_jst
        import os as _os
        api_key = _os.environ.get("GEMINI_API_KEY")
        if api_key:
            genai.configure(api_key=api_key)
            today_str = _now_jst().strftime("%Y-%m-%d (%A)")
            prompt = _build_pdf_native_prompt(today_str, file.filename or "uploaded.pdf")
            model_name = DEFAULT_MODELS.get("gemini", "gemini-2.0-flash-exp")
            mdl = genai.GenerativeModel(model_name=model_name)
            resp = mdl.generate_content(
                [
                    {"mime_type": "application/pdf", "data": contents},
                    prompt,
                ],
                generation_config={"max_output_tokens": 8192, "temperature": 0.2},
            )
            # Gemini may return empty if blocked/exceeded; access text defensively
            raw = ""
            try:
                raw = resp.text or ""
            except Exception:
                # resp.text raises when no candidates; surface diagnostic
                raw = ""
            if not raw:
                # Try to read from candidates / parts directly
                try:
                    parts = []
                    for c in getattr(resp, "candidates", []) or []:
                        for p in (getattr(c, "content", None) and c.content.parts) or []:
                            if getattr(p, "text", None):
                                parts.append(p.text)
                    raw = "".join(parts)
                except Exception:
                    pass
            if not raw:
                # Surface the reason
                pf = getattr(resp, "prompt_feedback", None)
                finish_reasons = []
                try:
                    for c in getattr(resp, "candidates", []) or []:
                        finish_reasons.append(str(getattr(c, "finish_reason", "")))
                except Exception:
                    pass
                diag = f"empty Gemini response (prompt_feedback={pf}, finish_reasons={finish_reasons})"
                # Fall through to PyPDF/pdfplumber text path below
                raise RuntimeError(diag)
            normalized, parse_error = _parse_proposals_json(raw, file.filename or "PDF")
            return {
                "proposals": normalized,
                "filename": file.filename,
                "page_count": 0,  # not extracted in this path
                "engine_used": "gemini-native-pdf",
                "model_used": model_name,
                "parse_error": parse_error,
                "ai_raw_preview": raw[:1500] if not normalized else None,
                "text_preview": "(Gemini が PDF を直接解析しました)",
                "text_length": len(contents),
            }
    except Exception as e:
        # fall through to text extraction path
        gemini_native_error = str(e)
    else:
        gemini_native_error = None

    text_pages: list[str] = []
    # First try pdfplumber (preserves tables/columns), fall back to PyPDF2
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(contents)) as pdf:
            for page in pdf.pages:
                page_parts: list[str] = []
                # extract tables separately so rows stay aligned
                try:
                    tables = page.extract_tables() or []
                    for table in tables:
                        for row in table:
                            cells = [str(c or "").strip() for c in row]
                            if any(cells):
                                page_parts.append(" | ".join(cells))
                except Exception:
                    pass
                # then plain text (avoid duplicating table content by keeping text last and unique-ish)
                try:
                    txt = page.extract_text() or ""
                    if txt.strip():
                        page_parts.append(txt)
                except Exception:
                    pass
                text_pages.append("\n".join(page_parts))
    except Exception:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(contents))
            for p in reader.pages:
                try:
                    text_pages.append(p.extract_text() or "")
                except Exception:
                    text_pages.append("")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"PDF parse failed: {e}")

    full_text = "\n\n---PAGE---\n\n".join(text_pages)

    if not full_text.strip():
        return {
            "proposals": [],
            "parse_error": "PDF からテキストを抽出できませんでした（画像のみ・スキャンPDF の可能性）",
            "filename": file.filename,
            "page_count": len(text_pages),
            "text_preview": "",
            "text_length": 0,
        }

    result = _extract_with_ai(full_text, file.filename or "uploaded.pdf", engine)
    result["filename"] = file.filename
    result["page_count"] = len(text_pages)
    result["text_preview"] = full_text[:3000]
    result["text_length"] = len(full_text)
    return result


@router.get("/calendar/upcoming")
def calendar_upcoming(days_ahead: int = Query(7, ge=1, le=180)):
    """Read upcoming events from Google Calendar (the source of truth)."""
    if not is_configured():
        raise HTTPException(status_code=400, detail="Google integration not configured")
    try:
        events = list_upcoming_events(days_ahead=days_ahead)
        return {"events": events, "count": len(events)}
    except HTTPException:
        raise
    except Exception as e:
        _raise_gcal_error(e, "Calendar read failed")


@router.get("/calendar/range")
def calendar_range(start: str = Query(...), end: str = Query(...)):
    """Read events between two ISO dates (YYYY-MM-DD). Used by the month grid."""
    if not is_configured():
        raise HTTPException(status_code=400, detail="Google integration not configured")
    try:
        from gcal import list_events_range
        events = list_events_range(start_date=start, end_date=end)
        return {"events": events, "count": len(events)}
    except HTTPException:
        raise
    except Exception as e:
        _raise_gcal_error(e, "Calendar read failed")


@router.delete("/calendar/event/{event_id}")
def calendar_delete_event(event_id: str, calendar_id: str = Query("primary"), slot: int = Query(1)):
    if not is_configured():
        raise HTTPException(status_code=400, detail="Google integration not configured")
    try:
        from gcal import delete_event
        delete_event(event_id, calendar_id=calendar_id, slot=slot)
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        _raise_gcal_error(e, "Calendar delete failed")


class ShiftReq(BaseModel):
    days: int  # +N で後ろにずらす、-N で前倒し
    calendar_id: str = "primary"
    slot: int = 1


@router.post("/calendar/event/{event_id}/shift")
def calendar_shift_event(event_id: str, body: ShiftReq):
    """予定を ±N 日ずらす (start/end を維持して日付だけスライド)."""
    if not is_configured():
        raise HTTPException(status_code=400, detail="Google integration not configured")
    try:
        from gcal import _get_service, update_event
        from datetime import datetime as _dt, timedelta as _td
        service = _get_service(body.slot)
        ev = service.events().get(calendarId=body.calendar_id, eventId=event_id).execute()
        is_all_day = "date" in (ev.get("start") or {})

        def shift(iso: str) -> str:
            if is_all_day:
                d = _dt.fromisoformat(iso[:10]) + _td(days=body.days)
                return d.strftime("%Y-%m-%d")
            # tz 保つ
            d = _dt.fromisoformat(iso.replace("Z", "+00:00")) + _td(days=body.days)
            return d.isoformat()

        new_start = shift(ev["start"].get("date") or ev["start"].get("dateTime"))
        new_end = shift(ev["end"].get("date") or ev["end"].get("dateTime"))
        result = update_event(
            event_id,
            calendar_id=body.calendar_id,
            slot=body.slot,
            start_iso=new_start,
            end_iso=new_end,
            all_day=is_all_day,
        )
        return {"ok": True, "id": result.get("id"), "shifted_days": body.days}
    except HTTPException:
        raise
    except Exception as e:
        _raise_gcal_error(e, "Calendar shift failed")


class EventPatch(BaseModel):
    title: str | None = None
    start_iso: str | None = None
    end_iso: str | None = None
    description: str | None = None
    location: str | None = None
    all_day: bool | None = None
    calendar_id: str = "primary"
    slot: int = 1


@router.patch("/calendar/event/{event_id}")
def calendar_patch_event(event_id: str, patch: EventPatch):
    if not is_configured():
        raise HTTPException(status_code=400, detail="Google integration not configured")
    try:
        from gcal import update_event
        result = update_event(
            event_id,
            calendar_id=patch.calendar_id,
            slot=patch.slot,
            title=patch.title,
            start_iso=patch.start_iso,
            end_iso=patch.end_iso,
            description=patch.description,
            location=patch.location,
            all_day=patch.all_day,
        )
        return {
            "ok": True,
            "id": result.get("id"),
            "html_link": result.get("htmlLink", ""),
            "summary": result.get("summary", ""),
        }
    except HTTPException:
        raise
    except Exception as e:
        _raise_gcal_error(e, "Calendar update failed")


class CreateEventRequest(BaseModel):
    title: str
    start_iso: str
    end_iso: str
    description: str = ""
    location: str = ""
    event_type: str | None = None  # meeting/committee/deadline/default; auto-detect if None
    recurrence: str = ""  # RRULE body, e.g. "FREQ=WEEKLY;UNTIL=20260801T235959Z"


@router.post("/calendar/create-event")
def create_event(req: CreateEventRequest):
    """Create a Google Calendar event with type-aware reminders and optional recurrence."""
    if not is_configured():
        raise HTTPException(status_code=400, detail="Google integration not configured")
    try:
        result = gcal_create_event(
            title=req.title,
            start_iso=req.start_iso,
            end_iso=req.end_iso,
            description=req.description,
            location=req.location,
            event_type=req.event_type,
            recurrence=req.recurrence or None,
        )
        return {
            "id": result.get("id", ""),
            "html_link": result.get("htmlLink", ""),
            "status": result.get("status", ""),
            "event_type_used": result.get("_event_type_used", "default"),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Calendar event creation failed: {e}")
